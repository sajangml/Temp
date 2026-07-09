from sys import argv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from netmiko import ConnectHandler, SSHDetect
from ntc_templates.parse import parse_output
from pprint import pprint
import argparse
import getpass
import json
import os
import platform
import logging
from ciscoconfparse2 import CiscoConfParse
import csv

redFill = PatternFill("solid", fgColor="DA9694")

logging.basicConfig(filename="debug.log",
                    format='%(asctime)s %(message)s',
                    filemode='w')

# Creating an object
logger = logging.getLogger(__name__)

# Setting the threshold of logger to DEBUG
logger.setLevel(logging.DEBUG)


def compare_int(interface1, interface2):
    int_tech1 = re.search(r'^[a-zA-Z]+', interface1)
    int_num1 = re.search(r'\d.*$', interface1)
    int_tech2 = re.search(r'^[a-zA-Z]+', interface2)
    int_num2 = re.search(r'\d.*$', interface2)
    if (int_tech1.group() in int_tech2.group() or int_tech2.group() in int_tech1.group()) and int_num1.group() == int_num2.group():
        return True
    else:
        return False

def delete_children_matching(parent, pattern):
    for child in reversed(list(parent.children)):
        if re.search(pattern, child.text):
            child.delete()

def process_ce(device):
    global userName
    global userPass

    full_arp = []
    vrf_interfaces = {}

    if not check_ping(device, 2, True):
        logger.debug('Device %s did not respond to ping. Skipping', device)
    else:
        try:
            detect = SSHDetect(host=device, port='22', username=userName, password=userPass, device_type='autodetect')
            logger.debug('Detected device type %s at %s', detect.autodetect(), device)
            device_type = detect.autodetect()
            connection = ConnectHandler(host=device, port='22', username=userName, password=userPass, device_type=device_type)
            if device_type=='cisco_xe':
                device_type='cisco_ios'
        except Exception as e:
            if 'name' in device_list[count]:
                logger.error('Had a problem connecting to %s %s.  %s', device, device_list[count]['name'], e)
            else:
                logger.error('Had a problem connecting to %s %s.  %s', device, device_list[count]['ip'], e)
        else:
            try:
                logger.debug('Processing %s', connection.find_prompt())
                ip_vrf_int_raw = connection.send_command('show ip vrf interfaces')
                version_raw = connection.send_command('show version')

            except Exception as e:
                logger.error('Something broke getting raw VRF data from %s %s', device, e)
            else:
                if device_type == 'cisco_ios':
                    try:
                        ip_vrf_int_parsed = parse_output(platform=device_type, command="show ip vrf interface", data=ip_vrf_int_raw)
                    except Exception as e:
                        logger.error('Something went wrong parsing %s %s', device, e)
                    else:
                        for interface in ip_vrf_int_parsed:
                            vrf = interface['vrf']
                            if vrf not in vrf_interfaces:
                                logger.info('Adding VRF %s', vrf)
                                vrf_interfaces.update({vrf: []})
                            logger.info('Added interface %s to VRF %s', interface['interface'], vrf)
                            vrf_interfaces[vrf].append({'interface': interface['interface'], 'state': interface['proto_state']})
                else:
                    logger.debug('Device %s was identified as %s.', device, device_type)
            for vrf in vrf_interfaces:
                if vrf == 'Mgmt-vrf' or vrf == 'mgmtVrf':
                    pass
                else:
                    svi_up = False
                    for interface in vrf_interfaces[vrf]:
                        logger.info('Processing interface %s in VRF %s', interface['interface'], vrf)
                        if re.match('^Vl\d+', interface['interface']):
                            logger.info('Interface %s is a Vlan interface', interface['interface'])
                            if re.match('^Vl1[0-3]\d\d', interface['interface']):
                                logger.info('Interface %s is an uplink SVI', interface['interface'])
                            else:
                                logger.info('Interface %s is a user SVI', interface['interface'])
                                if interface['state'] == 'up':
                                    logger.info('Interface %s is up.  Get ARP data.', interface['interface'])
                                    svi_up = True
                                    show_command = 'show ip arp vrf ' + vrf + ' ' + interface['interface']
                                    ip_arp_raw = connection.send_command(show_command)
                                    ip_arp_parsed = parse_output(platform=device_type, command="show ip arp", data=ip_arp_raw)
                                    full_arp.extend(ip_arp_parsed)
                    if svi_up:
                        logger.info('VRF %s has at least one up user facing SVI', vrf)
                    else:
                        logger.info('VRF %s has no up user facing SVI - maybe delete.', vrf)

        return(full_arp,vrf_interfaces)

def process_device(switches):
    global userName
    global userPass
    global xlsx

    device = switches['switch']

    vrf_interfaces = {}
    isCe = False
    full_arp = None

    if not check_ping(device, 2, True):
        logger.debug('Device %s did not respond to ping. Skipping', device)
    else:
        try:
            detect = SSHDetect(host=device, port='22', username=userName, password=userPass, device_type='autodetect')
            logger.debug('Detected device type %s at %s', detect.autodetect(), device)
            device_type = detect.autodetect()
            connection = ConnectHandler(host=device, port='22', username=userName, password=userPass, device_type=device_type)
            if device_type=='cisco_xe':
                device_type='cisco_ios'
        except Exception as e:
            if 'name' in device_list[count]:
                logger.error('Had a problem connecting to %s %s.  %s', device, device_list[count]['name'], e)
            else:
                logger.error('Had a problem connecting to %s %s.  %s', device, device_list[count]['ip'], e)
        else:
            try:
                logger.debug('Processing %s', connection.find_prompt())
                ip_vrf_raw = connection.send_command('show ip vrf interface')
                version_raw = connection.send_command('show version')
                switchConfig = connection.send_command('show running-config')

            except Exception as e:
                logger.error('Something broke getting raw VRF data from %s %s', device, e)
            else:
                if device_type == 'cisco_ios':
                    count = 0
                    try:
                        ip_vrf_int_parsed = parse_output(platform=device_type, command="show ip vrf interface", data=ip_vrf_raw)
                    except Exception as e:
                        logger.error('Something went wrong parsing %s %s', device, e)
                    else:
                        for interface in ip_vrf_int_parsed:
                            vrf = interface['vrf']
                            if vrf not in vrf_interfaces:
                                logger.info('Adding VRF %s', vrf)
                                vrf_interfaces.update({vrf: []})
                            logger.info('Added interface %s to VRF %s', interface['interface'], vrf)
                            vrf_interfaces[vrf].append({'interface': interface['interface'], 'state': interface['proto_state']})   
                        for vrf in vrf_interfaces:
                            if vrf == 'Mgmt-vrf' or vrf == 'mgmtVrf':
                                pass
                            else:
                                count += 1
                                isCe = True

                    if not count:
                        logger.debug('Device %s seems to be an L2 switch.  Finding the CE', device)
                        ip_route_raw = connection.send_command('show ip route')
                        ip_route_parsed = parse_output(platform=device_type, command="show ip route", data=ip_route_raw)
                        if ip_route_parsed[0]['default_gateway'] != "":
                            logger.debug('Device %s has default-gateway %s.', device, ip_route_parsed[0]['default_gateway'])
                            full_arp, vrf_interfaces = process_ce(ip_route_parsed[0]['default_gateway'])
                        else:
                            for route in ip_route_parsed:
                                if route['network'] == "0.0.0.0":
                                    logger.debug('Device %s has default route via %s.', device, route['nexthop_ip'])
                                    full_arp, vrf_interfaces = process_ce(route['nexthop_ip'])
                                    break
                                else:
                                    logger.debug('Route to %s is not default.', route['nexthop_ip'])
                    else:
                        logger.debug('Device %s seems to be a CE.', device)
                        full_arp, vrf_interfaces = process_ce(device)
                else:
                    logger.debug('Device %s was identified as %s.', device, device_type)

                        
            config_parsed = CiscoConfParse(switchConfig.splitlines(), syntax='ios', factory=False, auto_commit=True)
            int_status_raw = connection.send_command('show interface status')
            int_status_parsed = parse_output(platform=device_type, command="show int status", data=int_status_raw)
            int_desc_raw = connection.send_command('show interface description')
            int_desc_parsed = parse_output(platform=device_type, command="show int desc", data=int_desc_raw)
            mac_table_raw = connection.send_command('show mac address-table')
            mac_table_parsed = parse_output(platform=device_type, command="show mac address-table", data=mac_table_raw)
            vlan_raw = connection.send_command('show vlan')
            vlan_parsed = parse_output(platform=device_type, command="show vlan", data=vlan_raw)
            version_parsed = parse_output(platform=device_type, command="show version", data=version_raw)
            cdpRaw = connection.send_command('show cdp neighbors detail')
            cdpParsed = parse_output(platform=device_type, command="show cdp neighbors detail", data=cdpRaw)
            ws = xlsx.create_sheet('Port Status',0)

            if switches['prefile'] is None:
                with open(version_parsed[0]['hostname']+'-prechange.json', 'w') as prechangeFile:
                    prechangeFile.write(json.dumps(int_status_parsed))
                    prechangeFile.write(json.dumps(int_desc_parsed))
                    prechangeFile.write(json.dumps(mac_table_parsed))
                    prechangeFile.write(json.dumps(vlan_parsed))
                    prechangeFile.write(json.dumps(version_parsed))
                    prechangeFile.write(json.dumps(cdpParsed))
                    if full_arp:
                        prechangeFile.write(json.dumps(full_arp))
            else:
                with open(switches['prefile']) as prechangeFile:
                    preJson = prechangeFile.read()
                    prePortDict = json.loads(preJson)
                pprint(prePortDict)

            if True:
                ws.append(['Port', 'Description', 'Status', 'MAC', 'IP', 'VLAN', 'VLAN Name', 'Media Type', 'Speed', 'Duplex'])
                headerCellFont = Font(bold=True, color="FFFFFF")
                headerCellFill = PatternFill("solid", fgColor="1F497D")
                colRange = range(1,11)
                for col in colRange:
                    ws.cell(row=1, column=col).fill = headerCellFill
                    ws.cell(row=1, column=col).font = headerCellFont


                for int_status in int_status_parsed:
                    if 'Vl' in int_status['port'] or 'Tu' in int_status['port']:
                        logger.info('Not processing port %s', int_status['port'])
                    else:
                        logger.info('Processing port %s', int_status['port'])
                        for port in int_desc_parsed:
                            if port['port'] == int_status['port']:
                                break
                        port_dict = dict()
                        port_dict['port'] = port['port']
                        port_dict['desc'] = port['description']
                        port_dict['status'] = port['status']
                        port_dict['macs'] = []
                        port_mac = []

                        if int_status['vlan_id'] == "trunk":
                            logger.debug('Port %s is a trunk so writing a trunk entry', port_dict['port'])
                            ws.append([port_dict['port'], port_dict['desc'], port_dict['status'], '', '', 'trunk', '', int_status['type'], int_status['speed'], int_status['duplex']])
                            if int_status['speed'] == '10' or int_status['speed'] == 'a-10':
                                ws.cell(row=ws.max_row, column=9).fill = redFill
                            if int_status['duplex'] == 'half' or int_status['duplex'] == 'a-half':
                                ws.cell(row=ws.max_row, column=10).fill = redFill
                        if True:
                            for mac in mac_table_parsed:
                                if compare_int(mac['destination_port'][0], port['port']):
                                    logger.info('MAC address %s on port %s is the same as %s', mac['destination_address'], mac['destination_port'][0], port['port'])
                                    port_mac.append(mac)
                                else:
                                    logger.info('MAC address %s on port %s is not the same as %s', mac['destination_address'], mac['destination_port'][0], port['port'])
                            for mac in port_mac:
                                hasArp = False
                                for arp_entry in full_arp:
                                    if arp_entry['mac_address'] == mac['destination_address'] and arp_entry['interface'] == 'Vlan' + mac['vlan_id']:
                                        logger.info('MAC address %s has an ARP entry in VLAN %s' , mac['destination_address'], mac['vlan_id'])
                                        port_dict['macs'].append({'mac': mac['destination_address'], 'ip': arp_entry['ip_address'], 'vlan': mac['vlan_id']})
                                        hasArp = True
                                        break
                                if not hasArp:
                                    logger.info('MAC address %s has no ARP entry.', mac['destination_address'])
                                    port_dict['macs'].append({'mac': mac['destination_address'], 'ip': '', 'vlan': mac['vlan_id']})

                            if len(port_dict['macs']) == 0:
                                logger.debug('Port %s had no MACs so writing a blank entry', port_dict['port'])
                                vlan_name = ''
                                for vlan in vlan_parsed:
                                    if int_status['vlan_id'] == vlan['vlan_id']:
                                        vlan_name = vlan['vlan_name'] 
                                ws.append([port_dict['port'], port_dict['desc'], port_dict['status'], '', '', int_status['vlan_id'], vlan_name, int_status['type'], int_status['speed'], int_status['duplex']])
                                if int_status['speed'] == '10' or int_status['speed'] == 'a-10':
                                    ws.cell(row=ws.max_row, column=9).fill = redFill
                                if int_status['duplex'] == 'half' or int_status['duplex'] == 'a-half':
                                    ws.cell(row=ws.max_row, column=10).fill = redFill
                            else:
                                logger.debug('Port %s has MACs so look up VLAN and write', port_dict['port'])
                                for mac in port_dict['macs']:
                                    vlan_name = ''
                                    for vlan in vlan_parsed:
                                        if mac['vlan'] == vlan['vlan_id']:
                                            vlan_name = vlan['vlan_name'] 
                                    ws.append([port_dict['port'], port_dict['desc'], port_dict['status'], mac['mac'], mac['ip'], mac['vlan'], vlan_name, int_status['type'], int_status['speed'], int_status['duplex']])
                                    if int_status['speed'] == '10' or int_status['speed'] == 'a-10':
                                        ws.cell(row=ws.max_row, column=9).fill = redFill
                                    if int_status['duplex'] == 'half' or int_status['duplex'] == 'a-half':
                                        ws.cell(row=ws.max_row, column=10).fill = redFill

                pprint(switches)
                

                if isCe:
                    logger.info('Switch has VRFs so treat it like a CE')
                    vrfWs = xlsx.create_sheet('VRF Routing')
                    vrfWs.append(['VRF', 'Protocol', 'Routing Source', 'Last Heard'])
                    colRange = range(1,5)
                    for col in colRange:
                        vrfWs.cell(row=1, column=col).fill = headerCellFill
                        vrfWs.cell(row=1, column=col).font = headerCellFont
                    vrfs_raw = connection.send_command('show vrf')
                    vrfs_parsed = parse_output(platform=device_type, command="show vrf", data=vrfs_raw)
                    for vrf in vrfs_parsed:
                        if vrf['default_rd'] == '<not set>':
                            logger.debug('VRF %s does not have an RD Set.  Skip it.', vrf['name'])
                        else:
                            if len(vrf['interfaces']):
                                logger.debug('VRF %s has RD %s and has %s interfaces.  Check for good routing.', vrf['name'], vrf['default_rd'], str(len(vrf['interfaces'])))
                                vrfProtocolsRaw = connection.send_command('show ip protocols vrf ' + vrf['name'])
                                vrfProtocolsParsed = parse_output(platform=device_type, command="show ip protocols", data=vrfProtocolsRaw)
                                for vrfProtocol in vrfProtocolsParsed:
                                    for idx, routeSource in enumerate(vrfProtocol['routing_source']):
                                        vrfWs.append([vrf['name'], vrfProtocol['protocol'], routeSource, vrfProtocol['last_update'][idx]])

                poeWs = xlsx.create_sheet('Power over Ethernet')
                poeWs.append(['Interface', 'Admin Status', 'Operational Status', 'Power'])
                colRange = range(1,5)
                for col in colRange:
                    poeWs.cell(row=1, column=col).fill = headerCellFill
                    poeWs.cell(row=1, column=col).font = headerCellFont
                poeRaw = connection.send_command('show power inline')
                try:
                    poeParsed = parse_output(platform=device_type, command="show power inline", data=poeRaw)
                except Exception as e:
                    logger.error('Could not parse PoE output.  %s', e)
                else:
                    for interface in poeParsed:
                        poeWs.append([interface['interface'], interface['admin_status'], interface['operational_status'], interface['power']])

                cdpWs = xlsx.create_sheet('CDP Neighbors')
                cdpWs.append(['Local Interface', 'Neighbor Name', 'Neighbor Interface'])
                colRange = range(1,4)
                for col in colRange:
                    cdpWs.cell(row=1, column=col).fill = headerCellFill
                    cdpWs.cell(row=1, column=col).font = headerCellFont
                
                for neighbor in cdpParsed:
                    cdpWs.append([neighbor['local_interface'], neighbor['neighbor_name'], neighbor['neighbor_interface']])

                configWs = xlsx.create_sheet('New Config')

                ethernetList = config_parsed.find_objects("^interface\s\w+thernet")
                
                portmap = {}
                with open('./portmap.csv', encoding='utf-8-sig', newline='') as csvfile:
                    portmapreader = csv.reader(csvfile)
                    for row in portmapreader:
                        portmap[row[0]] = row[1]

                deleteList = []
                for ethernet in ethernetList:
                    interfaceName = re.match(r'interface\s(\w+thernet.*)', ethernet.text).group(1)
                    if interfaceName in portmap:
                        ethernet.text = f'interface {portmap[interfaceName]}'
                    else:
                        logger.debug('No mapping for %s. Delete it.', interfaceName)
                        deleteList.append(ethernet)

                for intf in reversed(deleteList):
                    intf.delete()
                config_parsed.commit()

                for interface in reversed(config_parsed.find_objects("^interface\s\w+thernet")):
                    descriptionUpdated = False
                    name = re.match(r'interface\s(\w+thernet.*)', interface.text).group(1)
                    for neighbor in cdpParsed:
                        if not compare_int(name, neighbor['local_interface']):
                            logger.debug(f"Interface {name} does not match CDP local interface {neighbor['local_interface']}")
                            continue
                        pprint(neighbor['capabilities'].split(' '))
                        if not 'Switch' in neighbor['capabilities'].split(' '):
                            continue
                        delete_children_matching(interface, r'^\s+description')
                        shortNeighbor = neighbor['neighbor_name'].split('.')[0]
                        interface.append_to_family(f" description {shortNeighbor} {neighbor['neighbor_interface']}")
                        descriptionUpdated = True
                        print('Updated')
                        break
                    if not descriptionUpdated:
                        description = interface.re_search_children(r'\s+description\s---\s')
                        if description:
                            description[0].re_sub(r'\s+---\s+', ' ')
                            print(description[0].text)
                    config_parsed.commit()
                    currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                    delete_children_matching(currentInterface[0], r'^ mab')
                    config_parsed.commit()
                    currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                    delete_children_matching(currentInterface[0], r'^ dot1x')
                    config_parsed.commit()
                    currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                    delete_children_matching(currentInterface[0], r'^\sswitchport\sport-security\smac-address\ssticky\s')
                    config_parsed.commit()
                    currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                    if currentInterface[0].has_child_with(r'^ storm-control'):
                        delete_children_matching(currentInterface[0], r'^ storm-control')
                        config_parsed.commit()
                        currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                        currentInterface[0].append_to_family(' storm-control unknown-unicast level pps 10')
                        currentInterface[0].append_to_family(' storm-control multicast level bps 1m')
                        currentInterface[0].append_to_family(' storm-control broadcast level pps 10')
                        currentInterface[0].append_to_family(' service-policy output MQC-AutoQoS-1P7Q-Output-Policy')

                    config_parsed.commit()
                    currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                    if currentInterface[0].has_child_with(r'^ authentication'):
                        delete_children_matching(currentInterface[0], r'^ authentication')
                        config_parsed.commit()
                        currentInterface = config_parsed.find_objects(f'^interface\s{name}$')
                        currentInterface[0].append_to_family(' source template WIRED_DOT1X_CLOSED')

                config_parsed.commit()

                if isCe:
                    vrfConfigs = config_parsed.find_objects("^ip\svrf\s")
                    for vrfConfig in vrfConfigs:
                        configWs.append([str(vrfConfig.text)])
                        for child in vrfConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                if isCe:
                    mrouteConfigs = config_parsed.find_objects("^ip\smulticast-routing\s")
                    for mrouteConfig in mrouteConfigs:
                        configWs.append([str(mrouteConfig.text)])
                        for child in mrouteConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                stpConfigs = config_parsed.find_objects("^(no\s)?spanning-tree\s")
                for stpConfig in stpConfigs:
                    configWs.append([str(stpConfig.text)])
                    for child in stpConfig.all_children:
                        configWs.append([str(child.text)])
                    configWs.append(['!'])

                vlanConfigs = config_parsed.find_objects("^vlan\s\d+")
                for vlanConfig in vlanConfigs:
                    configWs.append([str(vlanConfig.text)])
                    for child in vlanConfig.all_children:
                        configWs.append([str(child.text)])
                    configWs.append(['!'])

                vlanFilterConfigs = config_parsed.find_objects("^vlan\sfilter")
                for vlanFilterConfig in vlanFilterConfigs:
                    configWs.append([str(vlanFilterConfig.text)])
                    for child in vlanFilterConfig.all_children:
                        configWs.append([str(child.text)])
                    configWs.append(['!'])

                interfaceConfigs = config_parsed.find_objects("^interface\s")
                for interfaceConfig in interfaceConfigs:
                    configWs.append([str(interfaceConfig.text)])
                    for child in interfaceConfig.all_children:
                        configWs.append([str(child.text)])
                    configWs.append(['!'])

                if isCe:
                    ripConfigs = config_parsed.find_objects("^router\srip")
                    for ripConfig in ripConfigs:
                        configWs.append([str(ripConfig.text)])
                        for child in ripConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                    ospfConfigs = config_parsed.find_objects("^router\sospf")
                    for ospfConfig in ospfConfigs:
                        configWs.append([str(ospfConfig.text)])
                        for child in ospfConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                    eigrpConfigs = config_parsed.find_objects("^router\seigrp")
                    for eigrpConfig in eigrpConfigs:
                        configWs.append([str(eigrpConfig.text)])
                        for child in eigrpConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                    pimConfigs = config_parsed.find_objects("^ip\s+pim")
                    for pimConfig in pimConfigs:
                        configWs.append([str(pimConfig.text)])
                        for child in pimConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                    staticRoutes = config_parsed.find_objects("^ip\s+route")
                    for staticRoute in staticRoutes:
                        configWs.append([str(staticRoute.text)])
                        configWs.append(['!'])

                    dhcpConfigs = config_parsed.find_objects("^ip\s+dhcp")
                    for dhcpConfig in dhcpConfigs:
                        configWs.append([str(dhcpConfig.text)])
                        for child in dhcpConfig.all_children:
                            configWs.append([str(child.text)])
                        configWs.append(['!'])

                sourceInterfaces = config_parsed.find_objects("source-interface")
                for sourceInterface in sourceInterfaces:
                    configWs.append([str(sourceInterface.text)])
                    configWs.append(['!'])

            return(version_parsed[0]['hostname'])


def check_ping(hostname, attempts = 1, silent = False):
    parameter = '-n' if platform.system().lower()=='windows' else '-c'
    filter = ' | findstr /i "TTL"' if platform.system().lower()=='windows' else ' | grep "ttl"'
    if (silent):
        silent = ' > NUL' if platform.system().lower()=='windows' else ' >/dev/null'
    else:
        silent = ''
    response = os.system('ping ' + parameter + ' ' + str(attempts) + ' ' + hostname + filter + silent)
    if response == 0:
        return True
    else:
        return False

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Switch LCM Data Collection", formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('--switch', action='store', required=True, help='IP address of the target switch')
    parser.add_argument('--prefile', action='store', required=False, help='JSON file from prechange run')
    args = parser.parse_args()
    switches = vars(args)

    userName = input('Username: ')
    userPass = getpass.getpass('Password: ')
#    with open('dump_list.txt') as file:
#        while line := file.readline():
#            switchAddress = line.rstrip()
    if True:
        if True:
            switchAddress = argv[1]
            xlsx = Workbook()
            switchName = process_device(switches)
            xlsx.save(switchName + "_Switch_Dump.xlsx")

